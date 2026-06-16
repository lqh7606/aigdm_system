import csv
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from pathlib import Path

from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime

from labs.models import LabResult
from labs.services import is_abnormal, normalize_lab_value
from maternal_records.models import MaternalRecord
from maternal_records.services import create_or_update_record_from_payload

from .models import ImportBatch, ImportTemplate, IntegrationSource, IntegrationTask


REQUIRED_IMPORT_FIELDS = {"record_no", "name"}
DEFAULT_MATERNAL_TEMPLATE_CODE = "MATERNAL_RECORD_IMPORT_V1"
DEFAULT_LAB_TEMPLATE_CODE = "LAB_RESULT_IMPORT_V1"

DEFAULT_IMPORT_TEMPLATE_COLUMNS = [
    {"field": "record_no", "title": "院内就诊号", "required": True},
    {"field": "name", "title": "姓名", "required": True},
    {"field": "id_card", "title": "证件号", "required": False},
    {"field": "phone", "title": "联系电话", "required": False},
    {"field": "age", "title": "年龄", "required": False},
    {"field": "last_menstrual_period", "title": "末次月经", "required": False},
    {"field": "expected_delivery_date", "title": "预产期", "required": False},
    {"field": "gestational_week", "title": "当前孕周", "required": False},
    {"field": "height_cm", "title": "身高cm", "required": False},
    {"field": "pre_preg_weight_kg", "title": "孕前体重kg", "required": False},
    {"field": "systolic_bp", "title": "收缩压", "required": False},
    {"field": "diastolic_bp", "title": "舒张压", "required": False},
    {"field": "diabetes_before_pregnancy", "title": "已确诊孕前糖尿病", "required": False},
]

DEFAULT_IMPORT_TEMPLATE_SAMPLE_ROWS = [
    {
        "院内就诊号": "P-SAMPLE-001",
        "姓名": "示例孕妇",
        "证件号": "110101199001011234",
        "联系电话": "13800000000",
        "年龄": 31,
        "末次月经": "2026-01-10",
        "预产期": "",
        "当前孕周": 10.5,
        "身高cm": 162,
        "孕前体重kg": 63.5,
        "收缩压": 118,
        "舒张压": 76,
        "已确诊孕前糖尿病": "否",
    }
]

DEFAULT_LAB_TEMPLATE_COLUMNS = [
    {"field": "record_no", "title": "院内就诊号", "required": True},
    {"field": "name", "title": "姓名", "required": False},
    {"field": "item_code", "title": "项目编码", "required": False},
    {"field": "item_name", "title": "项目名称", "required": False},
    {"field": "value", "title": "检验值", "required": True},
    {"field": "unit", "title": "单位", "required": False},
    {"field": "sampled_at", "title": "采样时间", "required": False},
    {"field": "reported_at", "title": "报告时间", "required": True},
]

DEFAULT_LAB_TEMPLATE_SAMPLE_ROWS = [
    {
        "院内就诊号": "P-SAMPLE-001",
        "姓名": "示例孕妇",
        "项目编码": "FPG",
        "项目名称": "空腹血糖",
        "检验值": "5.20",
        "单位": "mmol/L",
        "采样时间": "2026-03-20 08:00:00",
        "报告时间": "2026-03-20 10:00:00",
    }
]

IMPORT_HEADER_ALIASES = {
    "院内就诊号": "record_no",
    "档案编号": "record_no",
    "record_no": "record_no",
    "姓名": "name",
    "name": "name",
    "证件号": "id_card",
    "身份证号": "id_card",
    "id_card": "id_card",
    "联系电话": "phone",
    "手机号": "phone",
    "电话": "phone",
    "phone": "phone",
    "年龄": "age",
    "age": "age",
    "末次月经": "last_menstrual_period",
    "last_menstrual_period": "last_menstrual_period",
    "预产期": "expected_delivery_date",
    "expected_delivery_date": "expected_delivery_date",
    "当前孕周": "gestational_week",
    "孕周": "gestational_week",
    "gestational_week": "gestational_week",
    "身高cm": "height_cm",
    "身高": "height_cm",
    "height_cm": "height_cm",
    "孕前体重kg": "pre_preg_weight_kg",
    "孕前体重": "pre_preg_weight_kg",
    "pre_preg_weight_kg": "pre_preg_weight_kg",
    "收缩压": "systolic_bp",
    "systolic_bp": "systolic_bp",
    "舒张压": "diastolic_bp",
    "diastolic_bp": "diastolic_bp",
    "已确诊孕前糖尿病": "diabetes_before_pregnancy",
    "妊娠前糖尿病": "diabetes_before_pregnancy",
    "孕前糖尿病": "diabetes_before_pregnancy",
    "diabetes_before_pregnancy": "diabetes_before_pregnancy",
}

LAB_HEADER_ALIASES = {
    "院内就诊号": "record_no",
    "档案编号": "record_no",
    "record_no": "record_no",
    "姓名": "name",
    "name": "name",
    "项目编码": "item_code",
    "检验项目编码": "item_code",
    "item_code": "item_code",
    "项目名称": "item_name",
    "检验项目": "item_name",
    "检验项目名称": "item_name",
    "item_name": "item_name",
    "检验值": "value",
    "结果": "value",
    "value": "value",
    "单位": "unit",
    "unit": "unit",
    "采样时间": "sampled_at",
    "sampled_at": "sampled_at",
    "报告时间": "reported_at",
    "reported_at": "reported_at",
}

LAB_ITEM_NAME_TO_CODE = {
    "空腹血糖": "FPG",
    "FPG": "FPG",
    "OGTT 1小时血糖": "OGTT_1H",
    "OGTT1小时血糖": "OGTT_1H",
    "OGTT_1H": "OGTT_1H",
    "OGTT 2小时血糖": "OGTT_2H",
    "OGTT2小时血糖": "OGTT_2H",
    "OGTT_2H": "OGTT_2H",
    "甘油三酯": "TG",
    "TG": "TG",
    "高密度脂蛋白": "HDL_C",
    "HDL-C": "HDL_C",
    "HDL_C": "HDL_C",
}


class MockAdapter:
    def pull(self):
        return [
            {
                "record_no": "P-MOCK-001",
                "name": "王女士",
                "age": 31,
                "gestational_week": 10.5,
                "height_cm": 162,
                "pre_preg_weight_kg": 63.5,
                "pre_preg_bmi": 24.2,
                "last_menstrual_period": "2026-01-10",
            }
        ]


def normalize_import_kind(value):
    text = str(value or "").strip().upper()
    if text in {"LAB_RESULT", "LAB", "LABS", "检验数据"}:
        return ImportBatch.ImportKind.LAB_RESULT
    return ImportBatch.ImportKind.MATERNAL_RECORD


def _template_defaults(kind):
    kind = normalize_import_kind(kind)
    if kind == ImportBatch.ImportKind.LAB_RESULT:
        return {
            "code": DEFAULT_LAB_TEMPLATE_CODE,
            "name": "检验数据导入模板",
            "template_kind": ImportTemplate.TemplateKind.LAB_RESULT,
            "columns": DEFAULT_LAB_TEMPLATE_COLUMNS,
            "sample_rows": DEFAULT_LAB_TEMPLATE_SAMPLE_ROWS,
        }
    return {
        "code": DEFAULT_MATERNAL_TEMPLATE_CODE,
        "name": "孕产妇档案导入模板",
        "template_kind": ImportTemplate.TemplateKind.MATERNAL_RECORD,
        "columns": DEFAULT_IMPORT_TEMPLATE_COLUMNS,
        "sample_rows": DEFAULT_IMPORT_TEMPLATE_SAMPLE_ROWS,
    }


def get_or_create_default_import_template(kind=ImportBatch.ImportKind.MATERNAL_RECORD):
    defaults = _template_defaults(kind)
    template, created = ImportTemplate.objects.get_or_create(
        code=defaults["code"],
        defaults={
            "name": defaults["name"],
            "template_kind": defaults["template_kind"],
            "file_format": ImportTemplate.FileFormat.XLSX,
            "columns_json": defaults["columns"],
            "sample_rows_json": defaults["sample_rows"],
            "is_active": True,
        },
    )
    changed_fields = []
    if template.template_kind != defaults["template_kind"]:
        template.template_kind = defaults["template_kind"]
        changed_fields.append("template_kind")
    if not template.columns_json:
        template.columns_json = defaults["columns"]
        changed_fields.append("columns_json")
    if not template.sample_rows_json:
        template.sample_rows_json = defaults["sample_rows"]
        changed_fields.append("sample_rows_json")
    if changed_fields:
        changed_fields.append("updated_at")
        template.save(update_fields=changed_fields)
    return template


def get_active_import_template(kind=ImportBatch.ImportKind.MATERNAL_RECORD):
    kind = normalize_import_kind(kind)
    template_kind = (
        ImportTemplate.TemplateKind.LAB_RESULT
        if kind == ImportBatch.ImportKind.LAB_RESULT
        else ImportTemplate.TemplateKind.MATERNAL_RECORD
    )
    template = ImportTemplate.objects.filter(template_kind=template_kind, is_active=True).order_by("-updated_at").first()
    return template or get_or_create_default_import_template(kind)


def _normalize_boolean(value):
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return bool(value)
    return str(value).strip().lower() in {"是", "true", "1", "yes", "y", "on", "已确诊"}


def _normalize_row(row, aliases):
    normalized = {}
    for key, value in row.items():
        if key is None:
            continue
        clean_key = str(key).strip()
        target_key = aliases.get(clean_key, clean_key)
        if value == "":
            value = None
        normalized[target_key] = value
    return normalized


def normalize_import_row(row):
    normalized = _normalize_row(row, IMPORT_HEADER_ALIASES)
    if "diabetes_before_pregnancy" in normalized:
        normalized["diabetes_before_pregnancy"] = _normalize_boolean(normalized["diabetes_before_pregnancy"])
    return normalized


def normalize_import_rows(rows):
    return [normalize_import_row(row) for row in rows]


def normalize_lab_row(row):
    return _normalize_row(row, LAB_HEADER_ALIASES)


def normalize_lab_rows(rows):
    return [normalize_lab_row(row) for row in rows]


def _display_field_name(field):
    for item in DEFAULT_IMPORT_TEMPLATE_COLUMNS:
        if item["field"] == field:
            return item["title"]
    return field


def precheck_rows(rows):
    normalized_rows = normalize_import_rows(rows)
    errors = []
    for index, row in enumerate(normalized_rows, start=2):
        missing = sorted(field for field in REQUIRED_IMPORT_FIELDS if not row.get(field))
        if missing:
            missing_names = "、".join(_display_field_name(field) for field in missing)
            errors.append({"row": index, "rule": "REQ-ACCESS-003-C00", "message": f"缺少必填字段：{missing_names}"})
        if not row.get("last_menstrual_period") and not row.get("expected_delivery_date"):
            errors.append({"row": index, "rule": "REQ-ACCESS-003-C01", "message": "末次月经和预产期至少提供一项。"})
        height = row.get("height_cm")
        weight = row.get("pre_preg_weight_kg")
        if bool(height) != bool(weight):
            errors.append({"row": index, "rule": "REQ-ACCESS-003-C02", "message": "身高和孕前体重必须同时提供。"})
    return errors


def parse_csv_text(text):
    reader = csv.DictReader(StringIO(text))
    return [dict(row) for row in reader]


def parse_xlsx_file(file_obj):
    try:
        from openpyxl import load_workbook
    except ImportError:
        return parse_xlsx_file_stdlib(file_obj)
    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    parsed = []
    for row in rows[1:]:
        parsed.append({headers[index]: value for index, value in enumerate(row) if index < len(headers) and headers[index]})
    return parsed


def parse_xlsx_file_stdlib(file_obj):
    import re
    import zipfile
    import xml.etree.ElementTree as ET

    namespaces = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    with zipfile.ZipFile(file_obj) as archive:
        shared_strings = []
        if "xl/sharedStrings.xml" in archive.namelist():
            root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            for item in root.findall("a:si", namespaces):
                text = "".join(node.text or "" for node in item.findall(".//a:t", namespaces))
                shared_strings.append(text)

        sheet_names = [name for name in archive.namelist() if re.match(r"xl/worksheets/sheet\d+\.xml", name)]
        if not sheet_names:
            return []
        root = ET.fromstring(archive.read(sheet_names[0]))
        matrix = []
        for row in root.findall(".//a:sheetData/a:row", namespaces):
            values = []
            for cell in row.findall("a:c", namespaces):
                raw = cell.findtext("a:v", default="", namespaces=namespaces)
                if cell.attrib.get("t") == "s" and raw:
                    value = shared_strings[int(raw)]
                else:
                    value = raw
                values.append(value)
            matrix.append(values)
    if not matrix:
        return []
    headers = [str(item).strip() for item in matrix[0]]
    return [
        {headers[index]: value for index, value in enumerate(row) if index < len(headers) and headers[index]}
        for row in matrix[1:]
    ]


def parse_uploaded_file(uploaded_file):
    suffix = Path(uploaded_file.name).suffix.lower()
    if suffix == ".csv":
        content = uploaded_file.read()
        last_error = None
        for encoding in ("utf-8-sig", "gb18030"):
            try:
                return parse_csv_text(content.decode(encoding))
            except UnicodeDecodeError as exc:
                last_error = exc
        raise last_error
    if suffix == ".xlsx":
        return parse_xlsx_file(uploaded_file)
    raise ValueError("仅支持 CSV 或 XLSX 文件。")


def build_import_template_xlsx(template):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    defaults = _template_defaults(template.template_kind)
    columns = template.columns_json or defaults["columns"]
    sample_rows = template.sample_rows_json or defaults["sample_rows"]
    headers = [item.get("title") or item.get("field") for item in columns]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "导入模板"
    sheet.append(headers)
    for sample in sample_rows:
        sheet.append([sample.get(header, "") for header in headers])

    header_fill = PatternFill(fill_type="solid", fgColor="E2F3F0")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    sheet.freeze_panes = "A2"
    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 4, 12), 28)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_import_template_csv(template):
    defaults = _template_defaults(template.template_kind)
    columns = template.columns_json or defaults["columns"]
    sample_rows = template.sample_rows_json or defaults["sample_rows"]
    headers = [item.get("title") or item.get("field") for item in columns]
    buffer = StringIO()
    writer = csv.DictWriter(buffer, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    for sample in sample_rows:
        writer.writerow(sample)
    return "\ufeff" + buffer.getvalue()


def create_precheck_batch(file_name, rows, source=None):
    normalized_rows = normalize_import_rows(rows)
    errors = precheck_rows(normalized_rows)
    return ImportBatch.objects.create(
        source=source,
        import_kind=ImportBatch.ImportKind.MATERNAL_RECORD,
        file_name=file_name,
        status=ImportBatch.Status.FAILED if errors else ImportBatch.Status.PRECHECK,
        total_rows=len(normalized_rows),
        failed_rows=len(errors),
        error_json=errors,
    )


def import_rows(file_name, rows, source=None):
    normalized_rows = normalize_import_rows(rows)
    batch = create_precheck_batch(file_name, normalized_rows, source=source)
    if batch.error_json:
        return batch
    success = 0
    for row in normalized_rows:
        create_or_update_record_from_payload(row, source_type=MaternalRecord.SourceType.EXCEL, source_ref=file_name)
        success += 1
    batch.status = ImportBatch.Status.IMPORTED
    batch.success_rows = success
    batch.save(update_fields=["status", "success_rows", "updated_at"])
    return batch


def _parse_decimal(value):
    if value in (None, ""):
        raise ValueError("检验值不能为空。")
    try:
        return Decimal(str(value)).quantize(Decimal("0.001"))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError("检验值格式不正确。") from exc


def _parse_import_datetime(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = datetime.combine(value, time.min)
    else:
        text = str(value).strip()
        parsed = parse_datetime(text)
        if parsed is None:
            parsed_date = parse_date(text)
            parsed = datetime.combine(parsed_date, time.min) if parsed_date else None
    if parsed is None:
        return None
    if timezone.is_naive(parsed):
        parsed = timezone.make_aware(parsed, timezone.get_current_timezone())
    return parsed


def _resolve_lab_item(row):
    raw_code = str(row.get("item_code") or "").strip().upper()
    raw_name = str(row.get("item_name") or "").strip()
    if raw_code in LabResult.ItemCode.values:
        return raw_code, LabResult.ItemCode(raw_code).label
    if raw_name:
        mapped_code = LAB_ITEM_NAME_TO_CODE.get(raw_name) or LAB_ITEM_NAME_TO_CODE.get(raw_name.upper())
        if mapped_code:
            return mapped_code, LabResult.ItemCode(mapped_code).label
        for item_code, item_label in LabResult.ItemCode.choices:
            if raw_name == item_label:
                return item_code, item_label
    raise ValueError("项目编码无效，或项目名称无法识别。")


def _validate_lab_row(row, row_number):
    errors = []
    record_no = row.get("record_no")
    if not record_no:
        errors.append("缺少必填字段：院内就诊号")
    if not row.get("value"):
        errors.append("缺少必填字段：检验值")
    if not row.get("reported_at"):
        errors.append("缺少必填字段：报告时间")
    if not row.get("item_code") and not row.get("item_name"):
        errors.append("项目编码和项目名称至少提供一项")

    record = None
    if record_no:
        record = MaternalRecord.objects.filter(record_no=str(record_no).strip()).first()
        if record is None:
            errors.append(f"院内就诊号不存在：{record_no}")

    item_code = item_name = None
    if row.get("item_code") or row.get("item_name"):
        try:
            item_code, item_name = _resolve_lab_item(row)
        except ValueError as exc:
            errors.append(str(exc))

    value = None
    if row.get("value") not in (None, ""):
        try:
            value = _parse_decimal(row.get("value"))
        except ValueError as exc:
            errors.append(str(exc))

    sampled_at = _parse_import_datetime(row.get("sampled_at"))
    reported_at = _parse_import_datetime(row.get("reported_at"))
    if row.get("sampled_at") and sampled_at is None:
        errors.append("采样时间格式不正确。")
    if row.get("reported_at") and reported_at is None:
        errors.append("报告时间格式不正确。")

    if errors:
        return None, {"row": row_number, "rule": "REQ-ACCESS-LAB", "message": "；".join(errors)}

    return {
        "record": record,
        "item_code": item_code,
        "item_name": item_name,
        "value": value,
        "unit": row.get("unit") or "mmol/L",
        "sampled_at": sampled_at,
        "reported_at": reported_at,
    }, None


def import_lab_rows(file_name, rows, source=None):
    normalized_rows = normalize_lab_rows(rows)
    errors = []
    success = 0
    overwritten = 0

    for index, row in enumerate(normalized_rows, start=2):
        payload, row_error = _validate_lab_row(row, index)
        if row_error:
            errors.append(row_error)
            continue

        abnormal = is_abnormal(payload["item_code"], payload["value"])
        confirmation_status = (
            LabResult.ConfirmationStatus.PENDING if abnormal else LabResult.ConfirmationStatus.CONFIRMED
        )
        confirmed_at = None if abnormal else timezone.now()

        result = LabResult.objects.filter(
            maternal_record=payload["record"],
            item_code=payload["item_code"],
            reported_at=payload["reported_at"],
        ).first()
        if result:
            overwritten += 1
            result.item_name = payload["item_name"]
            result.value = payload["value"]
            result.unit = payload["unit"]
            result.sampled_at = payload["sampled_at"]
            result.source_type = "EXCEL"
            result.source_ref = file_name
            result.is_abnormal = abnormal
            result.confirmation_status = confirmation_status
            result.confirmed_by = None
            result.confirmed_at = confirmed_at
            result.save(
                update_fields=[
                    "item_name",
                    "value",
                    "unit",
                    "sampled_at",
                    "source_type",
                    "source_ref",
                    "is_abnormal",
                    "confirmation_status",
                    "confirmed_by",
                    "confirmed_at",
                    "updated_at",
                ]
            )
        else:
            LabResult.objects.create(
                maternal_record=payload["record"],
                item_code=payload["item_code"],
                item_name=payload["item_name"],
                value=payload["value"],
                unit=payload["unit"],
                sampled_at=payload["sampled_at"],
                reported_at=payload["reported_at"],
                source_type="EXCEL",
                source_ref=file_name,
                is_abnormal=abnormal,
                confirmation_status=confirmation_status,
                confirmed_at=confirmed_at,
            )
        success += 1

    batch = ImportBatch.objects.create(
        source=source,
        import_kind=ImportBatch.ImportKind.LAB_RESULT,
        file_name=file_name,
        status=ImportBatch.Status.IMPORTED if success else ImportBatch.Status.FAILED,
        total_rows=len(normalized_rows),
        success_rows=success,
        failed_rows=len(errors),
        overwritten_rows=overwritten,
        error_json=errors,
    )
    return batch


def _validate_lab_row_v2(row, row_number):
    errors = []
    record_no = row.get("record_no")
    if not record_no:
        errors.append("missing record_no")
    if not row.get("value"):
        errors.append("missing value")
    if not row.get("reported_at"):
        errors.append("missing reported_at")
    if not row.get("item_code") and not row.get("item_name"):
        errors.append("missing item_code or item_name")

    record = None
    if record_no:
        record = MaternalRecord.objects.filter(record_no=str(record_no).strip()).first()
        if record is None:
            errors.append(f"record_no not found: {record_no}")

    item_code = item_name = None
    if row.get("item_code") or row.get("item_name"):
        try:
            item_code, item_name = _resolve_lab_item(row)
        except ValueError as exc:
            errors.append(str(exc))

    value = None
    if row.get("value") not in (None, ""):
        try:
            value = _parse_decimal(row.get("value"))
        except ValueError as exc:
            errors.append(str(exc))

    sampled_at = _parse_import_datetime(row.get("sampled_at"))
    reported_at = _parse_import_datetime(row.get("reported_at"))
    if row.get("sampled_at") and sampled_at is None:
        errors.append("invalid sampled_at")
    if row.get("reported_at") and reported_at is None:
        errors.append("invalid reported_at")

    unit = row.get("unit") or "mmol/L"
    standard_value = None
    standard_unit = unit
    threshold = None
    if not errors:
        try:
            standard_value, standard_unit, threshold = normalize_lab_value(
                item_code,
                value,
                unit=unit,
                department=getattr(record, "department", None),
            )
        except ValueError as exc:
            errors.append(str(exc))

    if errors:
        return None, {"row": row_number, "rule": "REQ-ACCESS-LAB", "message": "; ".join(errors)}

    return {
        "record": record,
        "item_code": item_code,
        "item_name": item_name,
        "value": value,
        "unit": unit,
        "standard_value": standard_value,
        "standard_unit": standard_unit,
        "threshold": threshold,
        "sampled_at": sampled_at,
        "reported_at": reported_at,
    }, None


def import_lab_rows(file_name, rows, source=None):
    normalized_rows = normalize_lab_rows(rows)
    batch = ImportBatch.objects.create(
        source=source,
        import_kind=ImportBatch.ImportKind.LAB_RESULT,
        file_name=file_name,
        status=ImportBatch.Status.PRECHECK,
        total_rows=len(normalized_rows),
        source_metadata_json={"entry": "excel_or_csv", "file_name": file_name},
    )
    errors = []
    success = 0
    overwritten = 0

    for index, row in enumerate(normalized_rows, start=2):
        payload, row_error = _validate_lab_row_v2(row, index)
        if row_error:
            errors.append(row_error)
            continue

        threshold = payload["threshold"]
        abnormal = is_abnormal(
            payload["item_code"],
            payload["value"],
            unit=payload["unit"],
            department=getattr(payload["record"], "department", None),
        )
        confirmation_status = (
            LabResult.ConfirmationStatus.PENDING if abnormal else LabResult.ConfirmationStatus.CONFIRMED
        )
        confirmed_at = None if abnormal else timezone.now()

        result = LabResult.objects.filter(
            maternal_record=payload["record"],
            item_code=payload["item_code"],
            reported_at=payload["reported_at"],
        ).first()
        if result:
            overwritten += 1
            result.item_name = payload["item_name"]
            result.value = payload["standard_value"]
            result.unit = payload["standard_unit"]
            result.raw_value = str(payload["value"])
            result.raw_unit = payload["unit"]
            result.standard_value = payload["standard_value"]
            result.standard_unit = payload["standard_unit"]
            result.sampled_at = payload["sampled_at"]
            result.source_type = "EXCEL"
            result.source_ref = file_name
            result.import_batch = batch
            result.is_abnormal = abnormal
            result.threshold_config = threshold["config"]
            result.threshold_snapshot_json = threshold["snapshot"]
            result.confirmation_status = confirmation_status
            result.confirmed_by = None
            result.confirmed_at = confirmed_at
            result.save(
                update_fields=[
                    "item_name",
                    "value",
                    "unit",
                    "raw_value",
                    "raw_unit",
                    "standard_value",
                    "standard_unit",
                    "sampled_at",
                    "source_type",
                    "source_ref",
                    "import_batch",
                    "is_abnormal",
                    "threshold_config",
                    "threshold_snapshot_json",
                    "confirmation_status",
                    "confirmed_by",
                    "confirmed_at",
                    "updated_at",
                ]
            )
        else:
            LabResult.objects.create(
                maternal_record=payload["record"],
                item_code=payload["item_code"],
                item_name=payload["item_name"],
                value=payload["standard_value"],
                unit=payload["standard_unit"],
                raw_value=str(payload["value"]),
                raw_unit=payload["unit"],
                standard_value=payload["standard_value"],
                standard_unit=payload["standard_unit"],
                sampled_at=payload["sampled_at"],
                reported_at=payload["reported_at"],
                source_type="EXCEL",
                source_ref=file_name,
                import_batch=batch,
                is_abnormal=abnormal,
                threshold_config=threshold["config"],
                threshold_snapshot_json=threshold["snapshot"],
                confirmation_status=confirmation_status,
                confirmed_at=confirmed_at,
            )
        success += 1

    batch.status = ImportBatch.Status.IMPORTED if success else ImportBatch.Status.FAILED
    batch.success_rows = success
    batch.failed_rows = len(errors)
    batch.overwritten_rows = overwritten
    batch.error_json = errors
    batch.save(update_fields=["status", "success_rows", "failed_rows", "overwritten_rows", "error_json", "updated_at"])
    return batch


def run_mock_pull(source):
    task = IntegrationTask.objects.create(source=source, status=IntegrationTask.Status.RUNNING, started_at=timezone.now())
    try:
        rows = MockAdapter().pull()
        for row in rows:
            create_or_update_record_from_payload(row, source_type=MaternalRecord.SourceType.HIS, source_ref=source.code)
        task.status = IntegrationTask.Status.SUCCESS
        task.pulled_count = len(rows)
        task.finished_at = timezone.now()
        task.save(update_fields=["status", "pulled_count", "finished_at"])
    except Exception as exc:
        task.status = IntegrationTask.Status.FAILED
        task.error_message = str(exc)
        task.finished_at = timezone.now()
        task.save(update_fields=["status", "error_message", "finished_at"])
    return task


def get_or_create_mock_source():
    source, _ = IntegrationSource.objects.get_or_create(
        code="MOCK_HIS",
        defaults={
            "name": "HIS/EMR 接入源",
            "source_kind": IntegrationSource.SourceKind.HIS,
            "adapter_path": "integrations.services.MockAdapter",
        },
    )
    if source.name.startswith("模拟"):
        source.name = "HIS/EMR 接入源"
        source.save(update_fields=["name", "updated_at"])
    changed_fields = []
    if not source.is_demo:
        source.is_demo = True
        changed_fields.append("is_demo")
    if not source.config_json.get("mode"):
        source.config_json = {**source.config_json, "mode": "demo_mock"}
        changed_fields.append("config_json")
    if changed_fields:
        changed_fields.append("updated_at")
        source.save(update_fields=changed_fields)
    return source
